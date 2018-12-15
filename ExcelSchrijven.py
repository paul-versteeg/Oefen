# Alle functies van OpenPyxl importeren
# OpenPyxl module kan je installeren met linux command: >> pip install openpyxl
# Daarna in PyCharm installeren door file->settings->project -> project interpreter. Op "+" klikken rechtsboven, naar openpyxl zoeken en installeren

from openpyxl import *

# excelsheet benoemen
wb=Workbook()
sheet=wb.active
# sheet=wb.create_sheet('eerste')
filepath="/home/dev/test.xlsx"

# een grove weinig elegante manier om wat in het excelfile te zetten
row=1
column=2

sheet.cell(row,column).value='kolom1'
row=row+1
sheet.cell(row,column).value=84
row=row+1
sheet.cell(row,column).value=39
row=row+1
sheet.cell(row,column).value=5
row=row+1
sheet.cell(row,column).value='=sum(b2:b4)' #formules moeten ALTIJD in het engels

column=column+1
row=1
sheet.cell(row,column).value='kolom2'
row=row+1
sheet.cell(row,column).value=230
row=row+1
sheet.cell(row,column).value=82
row=row+1
sheet.cell(row,column).value=9983
row=row+1
sheet.cell(row,column).value='=sum(c2:c4)' #formules moeten ALTIJD in het engels

column=column+1
row=1
sheet.cell(row,column).value='kolom3'
row=row+1
sheet.cell(row,column).value=864
row=row+1
sheet.cell(row,column).value=1
row=row+1
sheet.cell(row,column).value=33
row=row+1
sheet.cell(row,column).value='=sum(d2:d4)' #formules moeten ALTIJD in het engels

column=1
row=5
sheet.cell(row,column).value='Totaal'


# file opslaan
wb.save(filepath)

